import openpyxl as xl
# import charts
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    # cell = sheet['a1']
    # or
    # cell = sheet.cell(1, 1)
    # print(cell.value)

    # iterate all the row and columns
    print(sheet.max_row)

    for row in range(2, sheet.max_row + 1):
        # print(row)
        cell = sheet.cell(row, 3)
        # print(cell.value)
        # multiply the cell by 0.9
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price
        # save the value to the xlsx

    # select the column and row ranges for our chart
    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'E2')

    wb.save('filename')
